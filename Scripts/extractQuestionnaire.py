import openpyxl, os

from docx import Document

# Load excel file
xl = openpyxl.load_workbook('C:/Users/Karan/Downloads/Website_Questionnaire_Form.xlsx')

sheet = xl['Sheet1']

#columns and index
letters = ["B", "C", "D", "E", "F", "G", "H","I", "J", "AL", "AV", "AW", "AX", "AY", "AU", "AT", "AN", "AO", "AP", 
           "AQ", "AS", "AR", "AM", "AH", "AJ", "AK", "AI", "AB", "AC", "Y", "AD", "AE", "AF", "AA", "AG", "X", "W", 
           "N", "Z", "O", "P", "R", "Q", "V", "S", "T", "U", "K", "L", "M"]

remainder = (sheet.max_row - 1) % 2

maxRow = (((sheet.max_row - 1) // 2) + remainder)

answers = []

i = 0

# Parse through worksheet to manipulate cells
while i < len(letters):
    answers.append(sheet[letters[i] + str(2)].value)
    i+=1

print(answers)
